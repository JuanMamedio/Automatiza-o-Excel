import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from copy import copy

# ================= APARÊNCIA =================
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("dark-blue")

# ================= JANELA =================
janela = ctk.CTk()
janela.title("Split_Envios_Industrializacao_Iguacu_V1.12")
janela.geometry("380x420")
janela.resizable(False, False)

# ================= VARIÁVEIS =================
caminho_excel = None
aba_selecionada = tk.StringVar(master=janela)
coluna_grupo = tk.StringVar(master=janela)
coluna_valor = tk.StringVar(master=janela)

# ================= FUNÇÕES AUX =================
def copiar_celula_com_estilo(origem, destino):
    destino.value = origem.value
    destino.font = copy(origem.font)
    destino.fill = copy(origem.fill)
    destino.alignment = copy(origem.alignment)
    destino.number_format = origem.number_format
    destino.border = copy(origem.border)
    destino.protection = copy(origem.protection)

# ================= EXCEL =================
def selecionar_excel():
    global caminho_excel
    caminho_excel = filedialog.askopenfilename(
        title="Selecione o Excel",
        filetypes=[("Arquivos Excel", "*.xlsx")]
    )
    if caminho_excel:
        label_arquivo.configure(text=caminho_excel)
        carregar_abas()

def carregar_abas():
    wb = load_workbook(caminho_excel, data_only=True)
    abas = wb.sheetnames

    combo_abas.configure(values=abas)
    aba_selecionada.set(abas[0])
    combo_abas.set(abas[0])

    carregar_colunas()

def carregar_colunas():
    wb = load_workbook(caminho_excel, data_only=True)
    ws = wb[aba_selecionada.get()]

    headers = [
        ws.cell(1, c).value
        for c in range(1, ws.max_column + 1)
        if ws.cell(1, c).value
    ]

    combo_grupo.configure(values=headers)
    combo_valor.configure(values=headers)

    coluna_grupo.set(headers[0])
    coluna_valor.set(headers[0])

    combo_grupo.set(headers[0])
    combo_valor.set(headers[0])

# ================= PROCESSAMENTO =================
def processar():
    try:
        wb = load_workbook(caminho_excel)
        ws = wb[aba_selecionada.get()]

        grupo_col = valor_col = num_grupo_col = None

        for c in range(1, ws.max_column + 1):
            h = ws.cell(1, c).value
            if h == coluna_grupo.get():
                grupo_col = c
            if h == coluna_valor.get():
                valor_col = c
            if h == "NUM_GRUPO":
                num_grupo_col = c

        if not num_grupo_col:
            num_grupo_col = ws.max_column + 1
            ws.cell(1, num_grupo_col).value = "NUM_GRUPO"

        borda = Side(style="thin")
        numero_grupo = 1
        material_anterior = None
        linha_inicio = 2
        resumo = {}

        # ===== AGRUPAMENTO =====
        for linha in range(2, ws.max_row + 1):
            atual = ws.cell(linha, grupo_col).value

            if material_anterior and atual != material_anterior:
                for r in range(linha_inicio, linha):
                    for c in range(1, ws.max_column + 1):
                        cell = ws.cell(r, c)
                        cell.border = Border(
                            left=borda if c == 1 else cell.border.left,
                            right=borda if c == ws.max_column else cell.border.right,
                            top=borda if r == linha_inicio else cell.border.top,
                            bottom=borda if r == linha - 1 else cell.border.bottom
                        )
                numero_grupo += 1
                linha_inicio = linha

            ws.cell(linha, num_grupo_col).value = numero_grupo

            valor = ws.cell(linha, valor_col).value or 0
            resumo[numero_grupo] = resumo.get(numero_grupo, 0) + valor

            material_anterior = atual

        # ===== RESUMO =====
        if "RESUMO_GRUPOS" in wb.sheetnames:
            del wb["RESUMO_GRUPOS"]

        ws_res = wb.create_sheet("RESUMO_GRUPOS")
        ws_res.append(["GRUPO", "VALOR", "ACUMULADO", "ENVIO"])

        total = sum(resumo.values())
        limite1, limite2 = total / 3, total * 2 / 3

        acumulado = 0
        grupo_envio = {}

        for g in sorted(resumo):
            acumulado += resumo[g]
            envio = "Envio 1" if acumulado <= limite1 else "Envio 2" if acumulado <= limite2 else "Envio 3"
            grupo_envio[g] = envio
            ws_res.append([g, resumo[g], acumulado, envio])

        # ===== ENVIOS =====
        for envio in ["Envio 1", "Envio 2", "Envio 3"]:
            nome = envio.upper().replace(" ", "_")
            if nome in wb.sheetnames:
                del wb[nome]

            ws_env = wb.create_sheet(nome)

            for c in range(1, ws.max_column + 1):
                copiar_celula_com_estilo(ws.cell(1, c), ws_env.cell(1, c))

            linha_dest = 2
            for l in range(2, ws.max_row + 1):
                if grupo_envio.get(ws.cell(l, num_grupo_col).value) == envio:
                    for c in range(1, ws.max_column + 1):
                        copiar_celula_com_estilo(ws.cell(l, c), ws_env.cell(linha_dest, c))
                    linha_dest += 1

        wb.save(caminho_excel)
        messagebox.showinfo("Sucesso", "Processamento concluído com sucesso!")

    except Exception as e:
        messagebox.showerror("Erro", str(e))

# ================= INTERFACE =================
ctk.CTkLabel(
    janela,
    text="Split de Envios",
    font=ctk.CTkFont(size=15, weight="bold")
).pack(pady=8)

ctk.CTkButton(
    janela,
    text="Selecionar Arquivo",
    width=273.5,
    height=36,
    fg_color="#0E1FB8",
    hover_color="#203FEB",
    command=selecionar_excel
).pack(pady=6)

label_arquivo = ctk.CTkLabel(janela, text="Nenhum arquivo selecionado", wraplength=340)
label_arquivo.pack(pady=4)

frame = ctk.CTkFrame(janela)
frame.pack(pady=1, padx=54, fill="x")


# ===== AGRUPAR =====
ctk.CTkLabel(frame, text="Agrupar",text_color="#9B9B9B").grid(row=0, column=0, sticky="n", padx=6)
combo_grupo = ctk.CTkComboBox(frame, variable=coluna_grupo, width=260)
combo_grupo.grid(row=1, column=0, padx=6, pady=(0, 10))

# ===== VALOR =====
ctk.CTkLabel(frame, text="Valor",text_color="#9B9B9B").grid(row=2, column=0, sticky="n", padx=6)
combo_valor = ctk.CTkComboBox(frame, variable=coluna_valor, width=260)
combo_valor.grid(row=3, column=0, padx=6, pady=(0, 10))

# ===== ABA =====
ctk.CTkLabel(frame, text="Aba",text_color="#9B9B9B").grid(row=4, column=0, sticky="n", padx=6)
combo_abas = ctk.CTkComboBox(
    frame,
    variable=aba_selecionada,
    width=260,
    command=lambda _: carregar_colunas()
)
combo_abas.grid(row=5, column=0, padx=6)

# ===== PROCESSAR =====
ctk.CTkButton(
    janela,
    text="Processar",
    width=273.5,
    height=38,
    fg_color="#1B5E20",
    hover_color="#2E7D32",
    font=ctk.CTkFont(size=13, weight="bold"),
    command=processar
).pack(pady=20)

janela.mainloop()